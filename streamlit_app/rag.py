"""
RAG ligero usando TF-IDF (sin modelos pesados).
Funciona perfecto para volúmenes medianos (cientos de chunks).
"""

from io import BytesIO
import re
from typing import List, Tuple

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

import db


def extraer_texto_pdf(file_bytes: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(BytesIO(file_bytes))
    partes = []
    for page in reader.pages:
        try:
            t = page.extract_text() or ""
        except Exception:
            t = ""
        if t.strip():
            partes.append(t)
    return "\n\n".join(partes)


def extraer_texto_docx(file_bytes: bytes) -> str:
    from docx import Document
    doc = Document(BytesIO(file_bytes))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def extraer_texto_txt(file_bytes: bytes) -> str:
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            return file_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="ignore")

def extraer_texto_xlsx(file_bytes: bytes) -> str:
    """Extrae texto de un Excel: cada hoja como sección, filas como CSV legible."""
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    secciones = []
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        filas = []
        for fila in ws.iter_rows(values_only=True):
            celdas = [str(c) if c is not None else "" for c in fila]
            if any(c.strip() for c in celdas):
                filas.append(" | ".join(celdas))
        if filas:
            secciones.append(f"## Hoja: {nombre_hoja}\n" + "\n".join(filas))
    wb.close()
    return "\n\n".join(secciones)


def extraer_texto_csv(file_bytes: bytes) -> str:
    """Lee un CSV y lo presenta como texto tabular legible."""
    import csv
    texto_crudo = extraer_texto_txt(file_bytes)
    lector = csv.reader(texto_crudo.splitlines())
    filas = [" | ".join(c.strip() for c in fila) for fila in lector if any(c.strip() for c in fila)]
    return "\n".join(filas)


def extraer_texto(nombre_archivo: str, file_bytes: bytes) -> str:
    n = nombre_archivo.lower()
    if n.endswith(".pdf"):
        return extraer_texto_pdf(file_bytes)
    if n.endswith(".docx"):
        return extraer_texto_docx(file_bytes)
    if n.endswith(".xlsx") or n.endswith(".xlsm"):
        return extraer_texto_xlsx(file_bytes)
    if n.endswith(".csv"):
        return extraer_texto_csv(file_bytes)
    if n.endswith(".txt") or n.endswith(".md"):
        return extraer_texto_txt(file_bytes)
    raise ValueError(f"Formato no soportado: {nombre_archivo}")


def chunkear(texto: str, tam_chunk: int = 800, solapado: int = 100) -> List[str]:
    """Divide el texto en chunks por párrafos respetando el tamaño máximo."""
    texto = re.sub(r"\n{3,}", "\n\n", texto.strip())
    parrafos = [p.strip() for p in re.split(r"\n\n+", texto) if p.strip()]

    chunks: List[str] = []
    actual = ""
    for p in parrafos:
        if len(actual) + len(p) + 2 <= tam_chunk:
            actual = (actual + "\n\n" + p).strip() if actual else p
        else:
            if actual:
                chunks.append(actual)
            if len(p) <= tam_chunk:
                actual = p
            else:
                # Párrafo más grande que el chunk: dividir por palabras
                palabras = p.split()
                buf = []
                size = 0
                for w in palabras:
                    if size + len(w) + 1 > tam_chunk:
                        chunks.append(" ".join(buf))
                        buf = buf[-(solapado // 8):]  # solapado aproximado
                        size = sum(len(x) + 1 for x in buf)
                    buf.append(w)
                    size += len(w) + 1
                actual = " ".join(buf)
    if actual:
        chunks.append(actual)
    return chunks


def indexar_documento(agente: str, nombre_archivo: str, file_bytes: bytes,
                      subido_por: int) -> Tuple[int, int]:
    """Extrae texto, chunkea y guarda en BD. Retorna (doc_id, n_chunks)."""
    texto = extraer_texto(nombre_archivo, file_bytes)
    if not texto.strip():
        raise ValueError("El archivo está vacío o no se pudo extraer texto")
    chunks = chunkear(texto)
    if not chunks:
        raise ValueError("No se generaron fragmentos a partir del archivo")
    doc_id = db.guardar_documento(agente, nombre_archivo, texto, subido_por)
    db.guardar_chunks(doc_id, agente, chunks)
    return doc_id, len(chunks)


def recuperar_contexto(agente: str, consulta: str, top_k: int = 4) -> str:
    """Busca los chunks más relevantes para la consulta usando TF-IDF."""
    chunks = db.listar_chunks(agente)
    if not chunks:
        return ""

    textos = [c["texto"] for c in chunks]
    if not consulta.strip():
        return ""

    try:
        vect = TfidfVectorizer(
            lowercase=True,
            ngram_range=(1, 2),
            max_df=0.9,
            min_df=1,
        )
        matriz = vect.fit_transform(textos + [consulta])
        sim = cosine_similarity(matriz[-1], matriz[:-1])[0]
    except ValueError:
        return ""

    indices = sim.argsort()[::-1][:top_k]
    seleccionados = []
    for i in indices:
        if sim[i] <= 0:
            continue
        nombre = chunks[i].get("nombre_archivo") or "documento"
        seleccionados.append(f"[Fuente: {nombre}]\n{chunks[i]['texto']}")
    return "\n\n---\n\n".join(seleccionados)
